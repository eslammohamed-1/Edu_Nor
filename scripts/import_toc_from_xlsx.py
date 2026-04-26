#!/usr/bin/env python3
"""
استيراد جدول محتوى ToC من Excel إلى JSON متوافق مع نموذج Course في EduNor.
يربط العروض عبر curriculumOfferings.json عند التطابق (subjectSlug + stage + grade).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("ثبّت: pip install -r scripts/requirements-toc.txt", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parent.parent


def load_json(path: Path) -> Any:
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def norm_header(s: str | None) -> str:
    if s is None:
        return ""
    return str(s).strip()


def float_grade_to_index(g: Any) -> int:
    if g is None:
        return 1
    try:
        return int(float(g))
    except (TypeError, ValueError):
        return 1


def grade_to_stage_and_index(
    grade_num: int, stages_cfg: dict[str, list[str]]
) -> tuple[str, int]:
    """
    تحويل رقم الصف المطلق (1-12) إلى (مرحلة، فهرس نسبي).
    مثال: grade 7 → ("prep", 1) | grade 10 → ("secondary", 1)
    """
    primary_count = len(stages_cfg.get("primary", []))
    prep_count = len(stages_cfg.get("prep", []))

    if grade_num <= primary_count:
        return "primary", grade_num
    elif grade_num <= primary_count + prep_count:
        return "prep", grade_num - primary_count
    else:
        return "secondary", grade_num - primary_count - prep_count


def grade_label(stage: str, grade_num: int, stages: dict[str, list[str]]) -> str:
    labels = stages.get(stage) or []
    if 1 <= grade_num <= len(labels):
        return labels[grade_num - 1]
    return labels[0] if labels else f"grade-{grade_num}"


def stable_id(prefix: str, *parts: str) -> str:
    h = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:14]
    return f"{prefix}{h}"


def find_offering_id(
    offerings: list[dict],
    subject_slug: str,
    stage: str,
    grade: str,
) -> str | None:
    for o in offerings:
        if (
            o.get("subjectSlug") == subject_slug
            and o.get("stage") == stage
            and o.get("grade") == grade
            and not o.get("secondaryTrack")
        ):
            return o.get("id")
    for o in offerings:
        if (
            o.get("subjectSlug") == subject_slug
            and o.get("stage") == stage
            and o.get("grade") == grade
        ):
            return o.get("id")
    return None


def _safe_float(v: Any) -> float | None:
    """تحويل آمن لـ float — يرجع None للقيم غير الرقمية."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def rget(row: dict, *names: str) -> Any:
    for n in names:
        if n in row and row[n] is not None and str(row[n]).strip() != "":
            return row[n]
    for k, v in row.items():
        if not k:
            continue
        for n in names:
            if k.lower().replace("  ", " ").strip() == n.lower().strip():
                return v
    return None


def session_key(v: Any) -> str:
    s = cell_str(v)
    if not s or s == "None":
        return "1"
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


def lesson_type_from_row(r: dict, section_title: str) -> str:
    t = str(rget(r, "section_type", "Section Type", "Section Type ") or "").lower()
    st = str(rget(r, "session_type", "Session Type") or "").lower()
    comb = f"{t} {st} {section_title}".lower()
    if "quiz" in comb or "quizz" in comb:
        return "quiz"
    if "read" in comb or "revision" in comb:
        return "reading"
    return "video"


def build_courses(
    map_data: dict,
    offerings: list[dict],
    rows_by_key: list[tuple[tuple, list[dict]]],
) -> list[dict]:
    stages_cfg = map_data.get("stages", {})
    out: list[dict] = []
    for key, group_rows in rows_by_key:
        (subject_name, subj_id, stage, gidx, year, term, season) = key
        g_label = grade_label(stage, gidx, stages_cfg)
        offering_id = find_offering_id(offerings, subj_id, stage, g_label)
        by_session: dict[str, list[dict]] = defaultdict(list)
        for r in group_rows:
            sn = session_key(
                rget(
                    r,
                    "Session Number",
                    "session number",
                )
            )
            by_session[sn].append(r)

        chapters: list[dict] = []
        for si, skey in enumerate(sorted(by_session.keys(), key=lambda x: (len(x), x))):
            sess_rows = sorted(
                by_session[skey],
                key=lambda r: float_grade_to_index(
                    r.get("section number") or r.get("Section Number")
                ),
            )
            r0 = sess_rows[0] if sess_rows else {}
            sess_title = rget(r0, "Session Title", "session title") or f"وحدة {skey}"
            if not str(sess_title).strip():
                sess_title = f"وحدة {skey}"
            lessons: list[dict] = []
            for j, r in enumerate(sess_rows):
                sec_title = rget(r, "Section Title", "section title")
                if not str(sec_title or "").strip():
                    sec_title = f"جزء {j+1}"
                lid = stable_id("toc-l-", subj_id, stage, g_label, skey, str(j), str(sec_title))
                lt = lesson_type_from_row(
                    {str(k).lower().replace(" ", "_"): v for k, v in r.items()},
                    str(sec_title),
                )
                lessons.append(
                    {
                        "id": lid,
                        "title": str(sec_title)[:500],
                        "description": str(
                            r.get("notes")
                            or r.get("Notes")
                            or (sec_title)
                        )[:500],
                        "duration": 15,
                        "type": lt,
                        "content": f"<p>{str(sec_title)}</p>",
                        "order": j + 1,
                    }
                )
            if not lessons and sess_title:
                lessons.append(
                    {
                        "id": stable_id("toc-l-", subj_id, skey, "single"),
                        "title": str(sess_title)[:500],
                        "description": "",
                        "duration": 15,
                        "type": "video",
                        "content": f"<p>{sess_title}</p>",
                        "order": 1,
                    }
                )
            ch_id = stable_id("toc-ch-", subj_id, stage, g_label, skey, str(sess_title))
            chapters.append(
                {
                    "id": ch_id,
                    "title": str(sess_title)[:200],
                    "description": f"الجلسة {skey} — {year or ''}",
                    "order": si + 1,
                    "lessons": lessons,
                }
            )

        if not chapters:
            continue
        all_lessons = sum(len(c["lessons"]) for c in chapters)
        year_s = str(year) if year else "2025-2026"
        course_id = (
            f"toc-{subj_id}-{stage}-g{gidx}-" + hashlib.md5(repr(key).encode()).hexdigest()[:12]
        )
        course = {
            "id": course_id,
            "subjectId": subj_id,
            "title": f"{subject_name} — {g_label}",
            "description": f"جدول محتوى مستورد — العام {year_s} — {subject_name}",
            "stage": stage,
            "grade": g_label,
            "instructor": "منهج آلي (ToC)",
            "duration": int(all_lessons * 15),
            "lessonsCount": int(all_lessons),
            "studentsCount": 0,
            "rating": 0,
            "chapters": chapters,
            "academicYear": year_s,
            "term": _safe_float(term),
            "season": _safe_float(season),
        }
        if offering_id:
            course["offeringId"] = offering_id
        out.append(course)
    return out


def read_sheet_rows(
    ws,
) -> list[dict[str, Any]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = [norm_header(c) for c in header if c is not None]
    keys = [k.rstrip() for k in keys]
    out = []
    for row in rows_iter:
        if not any(x is not None and str(x).strip() for x in row):
            continue
        d = {}
        for i, k in enumerate(keys):
            if not k:
                continue
            if i < len(row):
                d[k] = row[i]
        out.append(d)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--input",
        type=Path,
        help="ملف .xlsx",
    )
    ap.add_argument(
        "--map",
        type=Path,
        default=ROOT / "app" / "src" / "config" / "tocImportMap.json",
    )
    ap.add_argument(
        "--offerings",
        type=Path,
        default=ROOT / "app" / "src" / "data" / "curriculumOfferings.json",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=ROOT / "app" / "src" / "data" / "toc" / "generated" / "coursesFromToc.json",
    )
    ap.add_argument(
        "--sheets",
        type=str,
        default="",
        help="قائمة أسماء مفصولة بفواصل; فارغ = كل الأوراق المسموحة",
    )
    ap.add_argument("--max-courses", type=int, default=0, help="0 = بلا حد")
    args = ap.parse_args()

    if not args.input or not args.input.exists():
        print("مطلوب --input يشير لملف xlsx", file=sys.stderr)
        return 1

    map_data = load_json(args.map)
    off_data = load_json(args.offerings)
    offerings: list[dict] = off_data.get("offerings", [])
    subject_names: dict = map_data.get("subjectNames", {})
    sheet_ctx: dict = map_data.get("sheetContext", {})
    skip: set = set(map_data.get("skipSheets", []))
    stages_cfg: dict = map_data.get("stages", {})

    wb = load_workbook(args.input, read_only=True, data_only=True)
    only_sheets = {s.strip() for s in args.sheets.split(",") if s.strip()}

    groups: dict[tuple, list[dict]] = defaultdict(list)
    for sheet in wb.worksheets:
        name = sheet.title.strip()
        if name in skip or name in ("",):
            continue
        if only_sheets and name not in only_sheets:
            continue

        ctx = sheet_ctx.get(name, {})
        sheet_stage = ctx.get("stage")  # قد يكون None إذا stageFromGrade=true
        use_grade_for_stage = ctx.get("stageFromGrade", False)

        # إذا لم يكن هناك stage ولا stageFromGrade، حاول الاستنتاج من الاسم
        if not sheet_stage and not use_grade_for_stage:
            low = name.lower()
            if "primary" in low:
                sheet_stage = "primary"
            elif "middle" in low:
                sheet_stage = "prep"
            elif any(
                name.startswith(p)
                for p in (
                    "Chemistry", "Biology", "Integrated",
                    "التاريخ", "الجغرافيا", "الفلسفة",
                )
            ):
                sheet_stage = "secondary"
            else:
                # لم نستطع تحديد المرحلة من الاسم — نستخدم الصف
                use_grade_for_stage = True
                print(
                    f"ℹ️  ورقة «{name}»: لم يُحدَّد stage — سيُستنتَج من رقم الصف",
                    file=sys.stderr,
                )

        data = read_sheet_rows(sheet)
        for r in data:
            subj = str(
                rget(r, "Subject", "Subject ", "subject", "subject_name") or ""
            ).strip()
            if not subj:
                continue
            subj_id = subject_names.get(subj)
            if not subj_id:
                print(
                    f"⚠️  مادة غير معرّفة: {subj!r} (ورقة «{name}»)",
                    file=sys.stderr,
                )
                continue

            g = float_grade_to_index(rget(r, "Grade", "grade"))

            # ── تحديد المرحلة والفهرس النسبي ──
            if use_grade_for_stage or not sheet_stage:
                stage, gidx = grade_to_stage_and_index(g, stages_cfg)
            else:
                stage = sheet_stage
                # تحويل الصف المطلق إلى نسبي داخل المرحلة
                max_g = len(stages_cfg.get(stage, []))
                if max_g and g > max_g:
                    # الصف أكبر من المرحلة المحددة — نستنتج المرحلة الصحيحة
                    stage, gidx = grade_to_stage_and_index(g, stages_cfg)
                else:
                    gidx = max(1, g)

            # ضمان أن gidx ضمن الحدود
            max_g = len(stages_cfg.get(stage, []))
            if max_g:
                gidx = max(1, min(gidx, max_g))

            year = rget(r, "Academic Year", "academic year") or "2025-2026"
            term = rget(r, "Term", "term")
            season = rget(r, "Season", "season")

            key = (subj, subj_id, stage, gidx, str(year), term, season)
            groups[key].append(r)
    wb.close()

    items = list(groups.items())
    if args.max_courses > 0:
        items = items[: args.max_courses]
    out_courses = build_courses(map_data, offerings, items)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as f:
        json.dump(out_courses, f, ensure_ascii=False, indent=2)
        f.write("\n")

    # ── ملخص ──
    subj_counts: dict[str, int] = defaultdict(int)
    for c in out_courses:
        subj_counts[c["subjectId"]] += 1
    print(f"\n✅ تمت كتابة {len(out_courses)} كورساً → {args.out}")
    print("─" * 50)
    for sid in sorted(subj_counts):
        print(f"   {sid:25s}  {subj_counts[sid]:3d} كورس")
    print("─" * 50)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
