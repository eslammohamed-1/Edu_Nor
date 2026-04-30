#!/usr/bin/env python3
"""
EduNor — Seed Script: يقرأ ملفات الإكسيل ويبني هيكل المنهج الكامل.
يدمج ملفين:
  1. Arabic & Social Sciences Index 2026-2027 (أحدث — أولوية للعربي والدراسات)
  2. 2025-2026 ToCs (Draft) (أشمل — باقي المواد)

يُصدّر: app/src/fixtures/demo-catalog/generated/curriculum.json
"""
import json
import os
import random
import sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# ──────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILE1 = os.path.expanduser("~/Downloads/Arabic & Social Sciences Index 2026 - 2027.xlsx")
FILE2 = os.path.expanduser("~/Downloads/2025-2026 ToCs (Draft).xlsx")
OUT_DIR = os.path.join(ROOT, "app", "src", "fixtures", "demo-catalog", "generated")

# ──────────────────────────────────────────────
# ID Generation
# ──────────────────────────────────────────────
_used_ids: set[str] = set()

def gen_id() -> str:
    while True:
        n = random.randint(100_000_000_000, 999_999_999_999)
        s = str(n)
        if s not in _used_ids:
            _used_ids.add(s)
            return s

# ──────────────────────────────────────────────
# Grade mapping: absolute grade → (stage_slug, grade_order, grade_name)
# ──────────────────────────────────────────────
GRADE_MAP: dict[int, tuple[str, int, str]] = {}

_primary_names = [
    "الصف الأول الابتدائي", "الصف الثاني الابتدائي", "الصف الثالث الابتدائي",
    "الصف الرابع الابتدائي", "الصف الخامس الابتدائي", "الصف السادس الابتدائي",
]
_prep_names = [
    "الصف الأول الإعدادي", "الصف الثاني الإعدادي", "الصف الثالث الإعدادي",
]
_sec_names = [
    "الصف الأول الثانوي", "الصف الثاني الثانوي", "الصف الثالث الثانوي",
]

for i, name in enumerate(_primary_names, 1):
    GRADE_MAP[i] = ("primary", i, name)
for i, name in enumerate(_prep_names, 1):
    GRADE_MAP[6 + i] = ("prep", i, name)
for i, name in enumerate(_sec_names, 1):
    GRADE_MAP[9 + i] = ("secondary", i, name)

# ──────────────────────────────────────────────
# Subject metadata
# ──────────────────────────────────────────────
SUBJECT_META: dict[str, dict] = {
    "اللغة العربية":           {"slug": "arabic",             "icon": "BookOpen",      "color": "#1E3A5F"},
    "اللغة الإنجليزية":        {"slug": "english",            "icon": "Languages",     "color": "#3498DB"},
    "English":                  {"slug": "english",            "icon": "Languages",     "color": "#3498DB"},
    "الرياضيات":               {"slug": "math",               "icon": "Calculator",    "color": "#F4A825"},
    "Mathematics":              {"slug": "math",               "icon": "Calculator",    "color": "#F4A825"},
    "Mathematics - Scientific Section": {"slug": "math-sci",   "icon": "Calculator",    "color": "#F4A825"},
    "Mathematics - Arts Section":       {"slug": "math-arts",  "icon": "Calculator",    "color": "#E67E22"},
    "Statistics":               {"slug": "statistics",         "icon": "BarChart3",     "color": "#2874A6"},
    "العلوم":                  {"slug": "science",            "icon": "Atom",          "color": "#2EC4B6"},
    "Science":                  {"slug": "science",            "icon": "Atom",          "color": "#2EC4B6"},
    "Integrated Science":       {"slug": "integrated-science", "icon": "FlaskConical",  "color": "#117A65"},
    "الدراسات الاجتماعية":      {"slug": "social-studies",     "icon": "LandPlot",      "color": "#C0392B"},
    "التاريخ":                 {"slug": "history",            "icon": "Scroll",        "color": "#A0522D"},
    "جغرافيا":                 {"slug": "geography",          "icon": "Globe2",        "color": "#1E8449"},
    "جغرافيا ":                {"slug": "geography",          "icon": "Globe2",        "color": "#1E8449"},  # trailing space
    "الفلسفة والمنطق":         {"slug": "philosophy",         "icon": "Lightbulb",     "color": "#7D6608"},
    "علم النفس والاجتماع":      {"slug": "psychology-social",  "icon": "Users",         "color": "#AF7AC5"},
    "Physics":                  {"slug": "physics",            "icon": "Zap",           "color": "#9B59B6"},
    "Chemistry":                {"slug": "chemistry",          "icon": "FlaskConical",  "color": "#27AE60"},
    "Biology":                  {"slug": "biology",            "icon": "Dna",           "color": "#E67E22"},
    "Connect Plus":             {"slug": "connect-plus",       "icon": "Link2",         "color": "#1ABC9C"},
    "Discover":                 {"slug": "discover",           "icon": "Compass",       "color": "#8E44AD"},
    "تكنولوجيا المعلومات":      {"slug": "ict",               "icon": "Laptop",        "color": "#16A085"},
}

def get_subject_meta(name: str) -> dict:
    name = name.strip()
    if name in SUBJECT_META:
        return SUBJECT_META[name]
    slug = name.lower().replace(" ", "-").replace("(", "").replace(")", "")
    return {"slug": slug, "icon": "BookOpen", "color": "#1E3A5F"}

# ──────────────────────────────────────────────
# Data structures
# ──────────────────────────────────────────────
# stage_slug → Stage dict
stages: dict[str, dict] = {}
# (stage_slug, grade_order) → Grade dict
grades: dict[tuple[str, int], dict] = {}
# (stage_slug, grade_order, term_order) → Term dict
terms: dict[tuple[str, int, int], dict] = {}
# (stage_slug, grade_order, term_order, subject_slug) → Subject dict
subjects: dict[tuple[str, int, int, str], dict] = {}
# Track which (subject_name, abs_grade, term) combos we've already loaded
loaded_combos: set[tuple[str, int, int]] = set()

def ensure_stage(slug: str) -> dict:
    if slug not in stages:
        name_map = {"primary": "ابتدائي", "prep": "إعدادي", "secondary": "ثانوي"}
        order_map = {"primary": 1, "prep": 2, "secondary": 3}
        stages[slug] = {
            "id": gen_id(),
            "name": name_map[slug],
            "slug": slug,
            "order": order_map[slug],
            "grades": [],
        }
    return stages[slug]

def ensure_grade(stage_slug: str, grade_order: int, grade_name: str) -> dict:
    key = (stage_slug, grade_order)
    if key not in grades:
        stage = ensure_stage(stage_slug)
        g = {
            "id": gen_id(),
            "stageId": stage["id"],
            "name": grade_name,
            "order": grade_order,
            "terms": [],
        }
        grades[key] = g
        stage["grades"].append(g)
    return grades[key]

def ensure_term(stage_slug: str, grade_order: int, grade_name: str, term_order: int) -> dict:
    key = (stage_slug, grade_order, term_order)
    if key not in terms:
        grade = ensure_grade(stage_slug, grade_order, grade_name)
        term_names = {1: "الترم الأول", 2: "الترم الثاني"}
        t = {
            "id": gen_id(),
            "gradeId": grade["id"],
            "name": term_names.get(term_order, f"ترم {term_order}"),
            "order": term_order,
            "track": None,
            "subjects": [],
        }
        terms[key] = t
        grade["terms"].append(t)
    return terms[key]

def ensure_subject(stage_slug: str, grade_order: int, grade_name: str,
                   term_order: int, subject_name: str) -> dict:
    meta = get_subject_meta(subject_name)
    slug = meta["slug"]
    key = (stage_slug, grade_order, term_order, slug)
    if key not in subjects:
        term = ensure_term(stage_slug, grade_order, grade_name, term_order)
        s = {
            "id": gen_id(),
            "termId": term["id"],
            "name": subject_name.strip(),
            "slug": slug,
            "icon": meta["icon"],
            "color": meta["color"],
            "order": len(term["subjects"]) + 1,
            "lessons": [],
        }
        subjects[key] = s
        term["subjects"].append(s)
    return subjects[key]

# ──────────────────────────────────────────────
# Parse a ToC sheet (standard format)
# Columns: Subject, Country, Language, Grade, Term, CourseType, AcademicYear,
#          Season, [MOE Lesson/]SessionNumber, SessionTitle, SessionType,
#          [SectionGroup,] SectionNumber, SectionTitle, SectionType, ...
# ──────────────────────────────────────────────
def parse_toc_sheet(ws, subject_override: str | None = None, only_regular: bool = True):
    """Parse a standard ToC sheet and add data to the global structures."""
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    # Find column indices
    def find_col(*names: str) -> int:
        for n in names:
            n_lower = n.lower()
            for i, h in enumerate(header):
                if n_lower in h:
                    return i
        return -1

    col_subject = find_col("subject")
    col_grade = find_col("grade")
    col_term = find_col("term")
    col_course_type = find_col("course type")
    col_session_num = find_col("session number")
    col_session_title = find_col("session title")
    col_session_type = find_col("session type")
    col_section_num = find_col("section number")
    col_section_title = find_col("section title")
    col_section_type = find_col("section type")
    col_section_group = find_col("section group", "moe l")  # MOE Lesson or Section Group = unit title
    col_has_quiz = find_col("has quiz", "has  question")
    col_has_worksheet = find_col("worksheet", "has worksheet")

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Subject
        subj_name = subject_override or (str(row[col_subject]).strip() if col_subject >= 0 and row[col_subject] else "")
        if not subj_name or subj_name == "None":
            continue

        # Grade (absolute: 1-12)
        raw_grade = row[col_grade] if col_grade >= 0 else None
        if not raw_grade:
            continue
        try:
            abs_grade = int(float(str(raw_grade)))
        except (ValueError, TypeError):
            continue
        if abs_grade not in GRADE_MAP:
            continue

        # Term
        raw_term = row[col_term] if col_term >= 0 else None
        try:
            term_order = int(float(str(raw_term))) if raw_term else 1
        except (ValueError, TypeError):
            term_order = 1
        if term_order < 1:
            term_order = 1

        # Course type filter
        if only_regular and col_course_type >= 0:
            ct = str(row[col_course_type]).strip() if row[col_course_type] else "Regular"
            if ct not in ("Regular", "None", ""):
                continue

        # Check if already loaded from a newer file
        combo = (get_subject_meta(subj_name)["slug"], abs_grade, term_order)
        if combo in loaded_combos:
            continue

        stage_slug, grade_order, grade_name = GRADE_MAP[abs_grade]

        # Session (= Lesson)
        session_num_raw = row[col_session_num] if col_session_num >= 0 else None
        try:
            session_num = int(float(str(session_num_raw))) if session_num_raw else 0
        except (ValueError, TypeError):
            session_num = 0

        session_title = str(row[col_session_title]).strip() if col_session_title >= 0 and row[col_session_title] else ""
        if not session_title or session_title == "None":
            continue

        session_type = str(row[col_session_type]).strip() if col_session_type >= 0 and row[col_session_type] else "Regular"
        if session_type == "None":
            session_type = "Regular"

        unit_title = None
        if col_section_group >= 0 and row[col_section_group]:
            ut = str(row[col_section_group]).strip()
            if ut and ut != "None":
                unit_title = ut

        # Section
        section_num_raw = row[col_section_num] if col_section_num >= 0 else None
        try:
            section_num = int(float(str(section_num_raw))) if section_num_raw else 1
        except (ValueError, TypeError):
            section_num = 1

        section_title = str(row[col_section_title]).strip() if col_section_title >= 0 and row[col_section_title] else ""
        if not section_title or section_title == "None":
            section_title = session_title  # fallback

        section_type = str(row[col_section_type]).strip() if col_section_type >= 0 and row[col_section_type] else "Regular"
        if section_type == "None":
            section_type = "Regular"

        has_quiz = False
        if col_has_quiz >= 0 and row[col_has_quiz]:
            hq = str(row[col_has_quiz]).strip().lower()
            has_quiz = hq in ("true", "1", "yes")

        has_ws = False
        if col_has_worksheet >= 0 and row[col_has_worksheet]:
            hw = str(row[col_has_worksheet]).strip().lower()
            has_ws = hw not in ("", "none", "false", "0", "no")

        # ── Build hierarchy ──
        subject = ensure_subject(stage_slug, grade_order, grade_name, term_order, subj_name)

        # Find or create lesson
        lesson = None
        for l in subject["lessons"]:
            if l["order"] == session_num and l["title"] == session_title:
                lesson = l
                break
        if not lesson:
            lesson = {
                "id": gen_id(),
                "subjectId": subject["id"],
                "title": session_title,
                "description": "",
                "duration": 0,
                "type": "video",
                "videoUrl": None,
                "content": "",
                "order": session_num,
                "unitTitle": unit_title,
                "sessionType": session_type,
                "sections": [],
            }
            subject["lessons"].append(lesson)
        elif unit_title and not lesson.get("unitTitle"):
            lesson["unitTitle"] = unit_title

        # Add section
        section = {
            "id": gen_id(),
            "lessonId": lesson["id"],
            "title": section_title,
            "type": section_type,
            "content": "",
            "duration": 0,
            "order": section_num,
            "hasQuiz": has_quiz,
            "hasWorksheet": has_ws,
        }
        lesson["sections"].append(section)


def mark_loaded(subject_slug: str, abs_grade: int, term_order: int):
    loaded_combos.add((subject_slug, abs_grade, term_order))


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────
def main():
    print("🔄 EduNor Seed: بدء دمج ملفات الإكسيل...")

    # ── Step 1: File 1 (2026-2027) — أحدث، أولوية ──
    if os.path.exists(FILE1):
        print(f"📄 قراءة الملف الأول (2026-2027)...")
        wb1 = openpyxl.load_workbook(FILE1, data_only=True)

        # Arabic RegularTOCs
        if "Arabic RegularTOCs" in wb1.sheetnames:
            ws = wb1["Arabic RegularTOCs"]
            parse_toc_sheet(ws, subject_override="اللغة العربية")
            # Mark Arabic grades from this file as loaded
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3]:
                    try:
                        g = int(float(str(row[3])))
                        t = int(float(str(row[4]))) if row[4] else 1
                        mark_loaded("arabic", g, t)
                    except: pass
            print(f"   ✅ اللغة العربية (2026-2027)")

        # Social Sciences Regular TOCs
        if " Social Sciences Regular TOCs" in wb1.sheetnames:
            ws = wb1[" Social Sciences Regular TOCs"]
            parse_toc_sheet(ws)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[3]:
                    try:
                        subj = str(row[0]).strip()
                        slug = get_subject_meta(subj)["slug"]
                        g = int(float(str(row[3])))
                        t = int(float(str(row[4]))) if row[4] else 1
                        mark_loaded(slug, g, t)
                    except: pass
            print(f"   ✅ الدراسات والفلسفة والجغرافيا وعلم النفس (2026-2027)")

        wb1.close()
    else:
        print(f"⚠️  الملف الأول غير موجود: {FILE1}")

    # ── Step 2: File 2 (2025-2026) — أشمل، يملأ الباقي ──
    if os.path.exists(FILE2):
        print(f"📄 قراءة الملف الثاني (2025-2026)...")
        wb2 = openpyxl.load_workbook(FILE2, data_only=True)

        # Content sheets to process
        content_sheets = {
            "اللغة العربية Term 1": None,
            "اللغة العربية Term 2": None,
            " Term 1 الدراسات الاجتماعية": None,
            "Term 2 الدراسات الاجتماعية": None,
            "الجغرافيا": None,
            "Term2 الجغرافيا": None,
            "التاريخ ترم اول": None,
            "التاريخ ترم ثاني": None,
            "الفلسفة والمنطق وعلم النفس والا": None,
            " term 2 الفلسفة والمنطق وعلم ال": None,
            "Discover": None,
            "Mathematics": None,
            "Primary Science": None,
            "Middle School Science": None,
            "Integrated Science": None,
            "Physics": None,
            "Biology": None,
            "Chemistry": None,
            "English": None,
            "Connect Plus": None,
        }

        for sheet_name in content_sheets:
            if sheet_name in wb2.sheetnames:
                ws = wb2[sheet_name]
                if ws.max_row < 2:
                    continue
                parse_toc_sheet(ws, only_regular=True)
                print(f"   ✅ {sheet_name}")

        wb2.close()
    else:
        print(f"⚠️  الملف الثاني غير موجود: {FILE2}")

    # ── Sort lessons within each subject ──
    for subj in subjects.values():
        subj["lessons"].sort(key=lambda l: l["order"])
        for lesson in subj["lessons"]:
            lesson["sections"].sort(key=lambda s: s["order"])

    # ── Build output ──
    output = {
        "stages": list(stages.values()),
        "meta": {
            "generatedAt": __import__("datetime").datetime.now().isoformat(),
            "totalStages": len(stages),
            "totalGrades": len(grades),
            "totalTerms": len(terms),
            "totalSubjects": len(subjects),
            "totalLessons": sum(len(s["lessons"]) for s in subjects.values()),
            "totalSections": sum(
                len(l["sections"])
                for s in subjects.values()
                for l in s["lessons"]
            ),
        },
    }

    # ── Write output ──
    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, "curriculum.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # ── Also write flat lists for Prisma seeding ──
    flat = {
        "stages": [],
        "grades": [],
        "terms": [],
        "subjects": [],
        "lessons": [],
        "sections": [],
    }
    for stage in stages.values():
        flat["stages"].append({k: v for k, v in stage.items() if k != "grades"})
        for grade in stage["grades"]:
            flat["grades"].append({k: v for k, v in grade.items() if k != "terms"})
            for term in grade["terms"]:
                flat["terms"].append({k: v for k, v in term.items() if k != "subjects"})
                for subj in term["subjects"]:
                    flat["subjects"].append({k: v for k, v in subj.items() if k != "lessons"})
                    for lesson in subj["lessons"]:
                        flat["lessons"].append({k: v for k, v in lesson.items() if k != "sections"})
                        for section in lesson["sections"]:
                            flat["sections"].append(section)

    flat_path = os.path.join(OUT_DIR, "curriculum_flat.json")
    with open(flat_path, "w", encoding="utf-8") as f:
        json.dump(flat, f, ensure_ascii=False, indent=2)

    # ── Print summary ──
    m = output["meta"]
    print()
    print("=" * 50)
    print(f"✅ تم بنجاح!")
    print(f"   المراحل:   {m['totalStages']}")
    print(f"   الصفوف:    {m['totalGrades']}")
    print(f"   الترمات:   {m['totalTerms']}")
    print(f"   المواد:    {m['totalSubjects']}")
    print(f"   الدروس:    {m['totalLessons']}")
    print(f"   السيكشنز:  {m['totalSections']}")
    print(f"   📁 {out_path}")
    print(f"   📁 {flat_path}")
    print("=" * 50)


if __name__ == "__main__":
    main()
